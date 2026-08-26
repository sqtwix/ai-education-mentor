import json
import re
from pathlib import Path
import openpyxl
import pypdf

BASE_DIR = Path(__file__).resolve().parent.parent
EXAMPLE_DIR = BASE_DIR / "example_files"
AI_DRIVER_DATA = BASE_DIR / "ai-driver" / "backend" / "data"
API_CORE_DATA = BASE_DIR / "api-core" / "ApiCore" / "ApiCore" / "Data"

AI_DRIVER_DATA.mkdir(parents=True, exist_ok=True)
API_CORE_DATA.mkdir(parents=True, exist_ok=True)

def build_catalog():
    courses = []
    seen_names = set()
    course_id_counter = 1

    # 1. Реестр электронных курсов (ЭК)
    ek_path = EXAMPLE_DIR / "Реестр электронных курсов.xlsx"
    if ek_path.exists():
        wb = openpyxl.load_workbook(ek_path, data_only=True)
        ws = wb.active
        for row_idx in range(2, ws.max_row + 1):
            name = ws.cell(row_idx, 1).value
            annot = ws.cell(row_idx, 2).value or ""
            target = ws.cell(row_idx, 3).value or ""
            results = ws.cell(row_idx, 4).value or ""
            
            if name and str(name).strip():
                clean_name = str(name).strip()
                if clean_name.lower() not in seen_names:
                    seen_names.add(clean_name.lower())
                    
                    courses.append({
                        "id": f"EK_{course_id_counter:03d}",
                        "name": clean_name,
                        "type": "ЭК",
                        "category": "Электронный курс",
                        "duration_hours": 0,
                        "annotation": str(annot).strip(),
                        "target": str(target).strip(),
                        "results": str(results).strip(),
                        "competencies": []
                    })
                    course_id_counter += 1

    # 2. Буклет ППК 2025. Названия, описания и часы берутся только из
    # карточек программ на страницах 11-87. Текст введения и произвольные
    # строки PDF не являются программами и не должны попадать в каталог.
    booklet_path = EXAMPLE_DIR / "Буклет Линейка программ на 2025 (2).pdf"
    if booklet_path.exists():
        reader = pypdf.PdfReader(booklet_path)
        category_ranges = [
            (11, 21, "Личностные компетенции"),
            (22, 30, "Инструменты управления"),
            (31, 69, "Профильные hard-skills"),
            (70, 76, "Профессиональные компетенции"),
            (77, 87, "Базовые компетенции"),
        ]

        for page_number in range(11, min(87, len(reader.pages)) + 1):
            page = reader.pages[page_number - 1]
            title_fragments = []
            text_fragments = []

            def collect_title(text, _cm, tm, _font, font_size):
                value = re.sub(r"\s+", " ", text).strip()
                if not value:
                    return
                y_position = float(tm[5])
                size = float(font_size)
                text_fragments.append((value, y_position, size))
                if 15.5 <= size <= 16.5:
                    title_fragments.append((value, y_position))

            page_text = page.extract_text(visitor_text=collect_title) or ""
            if not title_fragments:
                continue

            title_groups = []
            for fragment, y_position in title_fragments:
                if y_position <= 30 and title_groups:
                    title_groups[-1].append((fragment, y_position))
                    continue
                previous_positive_y = next(
                    (y for _, y in reversed(title_groups[-1]) if y > 30),
                    None
                ) if title_groups else None
                if not title_groups or previous_positive_y is None or abs(previous_positive_y - y_position) > 24:
                    title_groups.append([])
                title_groups[-1].append((fragment, y_position))

            page_hours = []
            for fragment_index, (fragment, y_position, font_size) in enumerate(text_fragments):
                if not 11.5 <= font_size <= 12.5 or "час" not in fragment.lower():
                    continue
                inline_numbers = re.findall(r"\d+", fragment)
                if inline_numbers and y_position > 30:
                    page_hours.append((y_position, int(inline_numbers[-1])))
                    continue

                previous_numeric = []
                previous_y = None
                for previous_fragment, candidate_y, candidate_size in reversed(text_fragments[:fragment_index]):
                    if not 11.5 <= candidate_size <= 12.5 or candidate_y <= 30:
                        continue
                    if not previous_fragment.isdigit():
                        if previous_numeric:
                            break
                        continue
                    if previous_y is None:
                        previous_y = candidate_y
                    if abs(candidate_y - previous_y) > 0.5:
                        break
                    previous_numeric.append(previous_fragment)
                if previous_numeric and previous_y is not None:
                    number_text = "".join(reversed(previous_numeric))
                    page_hours.append((previous_y, int(number_text)))

            category = next(
                label for start, end, label in category_ranges
                if start <= page_number <= end
            )

            for index, group in enumerate(title_groups):
                title_lines = [fragment for fragment, _ in group]
                official_title = " ".join(title_lines)
                normalized_title = re.sub(r"\s+", " ", official_title).strip()
                normalized_title = re.sub(r"САНКТ\s+-?\s*ПЕТЕРБУРГА", "САНКТ-ПЕТЕРБУРГА", normalized_title)
                if not normalized_title or normalized_title.lower() in seen_names:
                    continue

                start = page_text.find(title_lines[0])
                if start < 0:
                    continue
                body_start = start
                for title_line in title_lines:
                    line_position = page_text.find(title_line, body_start)
                    if line_position < 0:
                        break
                    body_start = line_position + len(title_line)

                body_end = len(page_text)
                if index + 1 < len(title_groups):
                    next_title = title_groups[index + 1][0][0]
                    next_position = page_text.find(next_title, body_start)
                    if next_position >= 0:
                        body_end = next_position

                block = page_text[body_start:body_end].strip()
                hours_match = re.search(r"(\d+)\s*ак\.\s*час(?:а|ов)?", block, re.IGNORECASE)
                key_topics_match = re.search(r"КЛЮЧЕВЫЕ\s+ТЕМЫ:", block, re.IGNORECASE)
                annotation_end_candidates = [
                    match.start() for match in (hours_match, key_topics_match) if match
                ]
                annotation_end = min(annotation_end_candidates) if annotation_end_candidates else len(block)

                positive_title_y = [y for _, y in group if y > 30]
                if not positive_title_y:
                    continue
                upper_y = min(positive_title_y)
                lower_title_edges = []
                for other_group in title_groups:
                    other_positive_y = [y for _, y in other_group if y > 30]
                    if other_positive_y and max(other_positive_y) < upper_y:
                        lower_title_edges.append(max(other_positive_y))
                lower_y = max(lower_title_edges, default=0)
                duration_candidates = [
                    (y, hours) for y, hours in page_hours if lower_y < y < upper_y
                ]
                if hours_match:
                    duration_hours = int(hours_match.group(1))
                elif duration_candidates:
                    _, duration_hours = max(duration_candidates, key=lambda item: item[0])
                else:
                    continue

                annotation = re.sub(r"\s+", " ", block[:annotation_end]).strip()
                if not annotation:
                    continue

                seen_names.add(normalized_title.lower())
                courses.append({
                    "id": f"PPK_{course_id_counter:03d}",
                    "name": normalized_title,
                    "type": "ППК",
                    "category": category,
                    "duration_hours": duration_hours,
                    "annotation": annotation,
                    "target": "",
                    "results": "",
                    "competencies": [category]
                })
                course_id_counter += 1

    return courses

def build_history_and_benchmarks():
    history_path = EXAMPLE_DIR / "выгрузка с историей обучения.xlsx"
    users_dict = {}
    total_records = 0
    
    if history_path.exists():
        wb = openpyxl.load_workbook(history_path, data_only=True)
        ws = wb["ГГС 2024-2025"]
        
        for r in range(2, ws.max_row + 1):
            fio = ws.cell(r, 1).value
            pos = ws.cell(r, 2).value
            dept = ws.cell(r, 3).value
            c_type = ws.cell(r, 4).value
            c_name = ws.cell(r, 5).value
            status = ws.cell(r, 6).value
            
            if fio and pos and c_name:
                total_records += 1
                fio_str = str(fio).strip()
                pos_str = str(pos).strip()
                dept_str = str(dept).strip() if dept else ""
                c_type_str = str(c_type).strip() if c_type else ""
                c_name_str = str(c_name).strip()
                status_str = str(status).strip()
                
                if fio_str not in users_dict:
                    users_dict[fio_str] = {
                        "fio": fio_str,
                        "position": pos_str,
                        "department": dept_str,
                        "experience_years": 0,
                        "career_goal": "",
                        "learning_history": []
                    }
                
                users_dict[fio_str]["learning_history"].append({
                    "course_name": c_name_str,
                    "course_type": c_type_str,
                    "status": status_str
                })

    # 1. Бенчмарки по должности (общегородские)
    pos_stats = {}
    for u in users_dict.values():
        p = u["position"]
        if p not in pos_stats:
            pos_stats[p] = {"total_employees": 0, "course_counts": {}, "course_passed": {}, "course_types": {}}
        pos_stats[p]["total_employees"] += 1
        
        for h in u["learning_history"]:
            cname = h["course_name"]
            ctype = h["course_type"]
            st = h["status"].lower()
            
            pos_stats[p]["course_counts"][cname] = pos_stats[p]["course_counts"].get(cname, 0) + 1
            pos_stats[p]["course_types"][cname] = ctype
            if st in ["пройден", "успешно", "passed", "done"]:
                pos_stats[p]["course_passed"][cname] = pos_stats[p]["course_passed"].get(cname, 0) + 1

    benchmarks_by_pos = {}
    for p, data in pos_stats.items():
        tot = data["total_employees"]
        c_bench = {}
        for cname, count in data["course_counts"].items():
            passed = data["course_passed"].get(cname, 0)
            c_bench[cname] = {
                "course_name": cname,
                "course_type": data["course_types"].get(cname, ""),
                "total_taken": count,
                "total_passed": passed,
                "popularity_pct": round((count / tot) * 100, 1) if tot > 0 else 0,
                "success_rate": round((passed / count) * 100, 1) if count > 0 else None
            }
        benchmarks_by_pos[p] = {
            "position": p,
            "total_employees": tot,
            "courses": c_bench
        }

    # 2. Бенчмарки по паре (Должность + ИОГВ)
    pos_dept_stats = {}
    for u in users_dict.values():
        pair_key = f"{u['position']}___{u['department']}"
        if pair_key not in pos_dept_stats:
            pos_dept_stats[pair_key] = {
                "position": u["position"],
                "department": u["department"],
                "total_employees": 0, 
                "course_counts": {}, 
                "course_passed": {}, 
                "course_types": {}
            }
        pos_dept_stats[pair_key]["total_employees"] += 1
        
        for h in u["learning_history"]:
            cname = h["course_name"]
            ctype = h["course_type"]
            st = h["status"].lower()
            
            pos_dept_stats[pair_key]["course_counts"][cname] = pos_dept_stats[pair_key]["course_counts"].get(cname, 0) + 1
            pos_dept_stats[pair_key]["course_types"][cname] = ctype
            if st in ["пройден", "успешно", "passed", "done"]:
                pos_dept_stats[pair_key]["course_passed"][cname] = pos_dept_stats[pair_key]["course_passed"].get(cname, 0) + 1

    benchmarks_by_pos_dept = {}
    for pair_key, data in pos_dept_stats.items():
        tot = data["total_employees"]
        c_bench = {}
        for cname, count in data["course_counts"].items():
            passed = data["course_passed"].get(cname, 0)
            c_bench[cname] = {
                "course_name": cname,
                "course_type": data["course_types"].get(cname, ""),
                "total_taken": count,
                "total_passed": passed,
                "popularity_pct": round((count / tot) * 100, 1) if tot > 0 else 0,
                "success_rate": round((passed / count) * 100, 1) if count > 0 else None
            }
        benchmarks_by_pos_dept[pair_key] = {
            "position": data["position"],
            "department": data["department"],
            "total_employees": tot,
            "courses": c_bench
        }

    return {
        "users": list(users_dict.values()),
        "total_records": total_records,
        "benchmarks_by_position": benchmarks_by_pos,
        "benchmarks_by_position_and_dept": benchmarks_by_pos_dept
    }

if __name__ == "__main__":
    print("1. Building clean catalog...")
    catalog = build_catalog()
    print(f"   --> Total catalog courses: {len(catalog)}")
    
    print("2. Building clean learning history & benchmarks...")
    history_data = build_history_and_benchmarks()
    print(f"   --> Total training records: {history_data['total_records']}")
    print(f"   --> Total users/profiles:   {len(history_data['users'])}")
    print(f"   --> Positions with benchmark: {len(history_data['benchmarks_by_position'])}")
    print(f"   --> Position+Dept pairs:      {len(history_data['benchmarks_by_position_and_dept'])}")

    # Сохраняем в оба сервиса
    for target_dir in [AI_DRIVER_DATA, API_CORE_DATA]:
        with open(target_dir / "courses_catalog.json", "w", encoding="utf-8") as f:
            json.dump(catalog, f, ensure_ascii=False, indent=2)
            
        with open(target_dir / "learning_history_dataset.json", "w", encoding="utf-8") as f:
            json.dump(history_data, f, ensure_ascii=False, indent=2)
            
    print("[SUCCESS] Datasets saved cleanly to ai-driver and api-core!")
